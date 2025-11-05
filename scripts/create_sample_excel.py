#!/usr/bin/env python3
"""
Create a sample Excel file for bulk upload testing
Requires: openpyxl
Install with: pip install openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import json
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Documents"

# Define headers
headers = ['title', 'body', 'tags', 'metadata']

# Style headers
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# Add headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Sample data
data = [
    {
        'title': 'Cybersecurity Fundamentals',
        'body': 'Cybersecurity is the practice of protecting systems, networks, and programs from digital attacks. These cyberattacks are usually aimed at accessing, changing, or destroying sensitive information.',
        'tags': 'security,cybersecurity,infosec',
        'metadata': json.dumps({'year': 2024, 'category': 'security', 'difficulty': 'beginner'})
    },
    {
        'title': 'Cloud Computing Architectures',
        'body': 'Cloud architecture is the way technology components combine to build a cloud, in which resources are pooled through virtualization technology and shared across a network.',
        'tags': 'cloud,architecture,infrastructure',
        'metadata': json.dumps({'year': 2024, 'category': 'cloud computing', 'difficulty': 'intermediate'})
    },
    {
        'title': 'DevOps Best Practices',
        'body': 'DevOps is a set of practices that combines software development and IT operations. It aims to shorten the systems development life cycle and provide continuous delivery with high software quality.',
        'tags': 'devops,ci/cd,automation',
        'metadata': json.dumps({'year': 2024, 'category': 'development', 'difficulty': 'intermediate'})
    },
    {
        'title': 'Microservices Architecture',
        'body': 'Microservices architecture is an approach to developing a single application as a suite of small services, each running in its own process and communicating with lightweight mechanisms.',
        'tags': 'microservices,architecture,distributed systems',
        'metadata': json.dumps({'year': 2024, 'category': 'architecture', 'difficulty': 'advanced'})
    },
    {
        'title': 'Data Science Workflow',
        'body': 'Data science is an interdisciplinary field that uses scientific methods, processes, algorithms and systems to extract knowledge and insights from structured and unstructured data.',
        'tags': 'data science,analytics,machine learning',
        'metadata': json.dumps({'year': 2024, 'category': 'data science', 'difficulty': 'intermediate'})
    },
    {
        'title': 'Agile Methodologies',
        'body': 'Agile software development comprises various approaches to software development under which requirements and solutions evolve through the collaborative effort of self-organizing and cross-functional teams.',
        'tags': 'agile,scrum,project management',
        'metadata': json.dumps({'year': 2024, 'category': 'methodology', 'difficulty': 'beginner'})
    },
    {
        'title': 'API Design Principles',
        'body': 'RESTful API design is an architectural style for designing networked applications. It relies on a stateless, client-server protocol, almost always HTTP.',
        'tags': 'api,rest,web services',
        'metadata': json.dumps({'year': 2024, 'category': 'development', 'difficulty': 'intermediate'})
    }
]

# Add data rows
for row_idx, item in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=item['title'])
    ws.cell(row=row_idx, column=2, value=item['body'])
    ws.cell(row=row_idx, column=3, value=item['tags'])
    ws.cell(row=row_idx, column=4, value=item['metadata'])

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 80
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 50

# Save the file
output_file = 'storage/app/sample_upload.xlsx'
wb.save(output_file)
print(f"✅ Sample Excel file created: {output_file}")
print(f"📊 Contains {len(data)} sample documents")
print("💡 Use this file to test the bulk upload feature")
